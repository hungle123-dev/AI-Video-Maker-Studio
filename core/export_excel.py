"""core/export_excel.py — Xuất danh sách bài của một project ra Excel (.xlsx).

Mỗi dòng = 1 bài, cột: STT · Tiêu đề · Mô tả · Hashtag · Đường dẫn video.
Dùng để chuẩn bị đăng YouTube/TikTok (copy tiêu đề/mô tả/hashtag, lấy file video).
"""
import os, re, unicodedata
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side; _STOP = set("là và của cho các một những với từ đến ở về theo như thì mà bằng khi trong ngoài trên dưới ra vào lên xuống này đó kia sẽ đã đang bị được có không cần tập bài phần chương số gì sự thật cách hai ba bốn năm sáu bảy tám chín mười the a an of for to in on and or is are be this that your you it".split()); _BRANDS = {"ai", "api", "gpt", "n8n", "nft", "seo", "gpt4", "gpt5", "grok", "canva", "excel", "figma", "react", "reels", "capcut", "claude", "crypto", "docker", "gemini", "notion", "openai", "python", "tiktok", "worker", "bitcoin", "chatgpt", "youtube", "cloudflare", "midjourney"}; _SUBJECT_TAGS = {"creator_economy": ["#KiemTienOnline", "#MMO", "#Freelance"], "tech": ["#CongNghe", "#AI", "#Tech"], "science": ["#KhoaHoc", "#TriThuc"], "general": ["#HocMoiNgay", "#KyNang"]}; _PREFIX = re.compile("^\\s*(t[aậ]p|b[aà]i|ph[aầ]n|ch[uươ]+ng|episode|ep|part)\\s*\\d+\\s*[:.\\-–—]?\\s*", re.IGNORECASE)
def _strip_accents(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s or "")
        if unicodedata.category(c) != "Mn"
    ).replace("đ", "d").replace("Đ", "D")

def make_hashtags(title: str, subject: str="", project_title: str="", limit: int=10) -> str:
    t = _PREFIX.sub("", title or ""); tokens = re.findall("[0-9A-Za-zÀ-ỹ]+", t); strong = []; content = []
    for w in tokens:
        wl = w.lower()
        has_digit = any((c.isdigit() for c in w))
        if wl in _STOP:
            continue
        if (has_digit and len(w) >= 2) or (w.isupper() and len(w) >= 2) or wl in _BRANDS:
            strong.append(w)
            continue
        if len(_strip_accents(w)) >= 2:
            content.append(w)
    tags = []; seen = set()
    def add(tag):
        k = tag.lower()
        if len(tag) > 2 and k not in seen:
            seen.add(k)
            tags.append(tag)
    for w in strong:
        s = _strip_accents(w)
        add("#" + (s if s.isupper() or any(c.isdigit() for c in s) else s.title()))
    if content:
        add("#" + "".join(_strip_accents(w).title() for w in content[:3]))
    if len(content) > 3:
        add("#" + "".join(_strip_accents(w).title() for w in content[3:6]))
    for b in _SUBJECT_TAGS.get(subject or "general", _SUBJECT_TAGS["general"]):
        add(b)
    return " ".join(tags[:limit])

def _safe_name(s: str) -> str:
    s = _strip_accents(s or "project"); s = re.sub("[^\\w\\s-]", "", s).strip().replace(" ", "_")
    return (s or "project")[:60]

def _xlsx_text(value: object) -> str:
    """Keep user/AI text as text instead of an Excel formula."""
    text = str(value or "")
    if text.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + text
    return text

def default_out_path(project_title: str) -> str:
    name = f"{_safe_name(project_title)}_danh_sach_bai.xlsx"
    for folder in (os.path.join(os.path.expanduser("~"), "Desktop"), os.path.join(os.path.expanduser("~"), "Downloads")):
        if os.path.isdir(folder):
            return os.path.join(folder, name)
    from config import DATA_DIR
    return os.path.join(str(DATA_DIR), name)


def _available_out_path(path: str) -> str:
    """Never silently replace an earlier export, including a default export."""
    candidate = os.path.abspath(path)
    if not os.path.exists(candidate):
        return candidate
    stem, extension = os.path.splitext(candidate)
    for number in range(2, 10_000):
        alternate = f"{stem} ({number}){extension}"
        if not os.path.exists(alternate):
            return alternate
    raise RuntimeError("Không tìm được tên file Excel trống để xuất.")

def export_project_xlsx(project_id: str, out_path: str=None) -> str:
    from core.project_store import project_store
    proj = project_store.get_project(project_id) or {}; lessons = project_store.list_lessons(project_id)
    if out_path is None:
        out_path = default_out_path(proj.get("title", "project"))
    out_path = _available_out_path(out_path)
    wb = Workbook(); ws = wb.active; ws.title = "Danh sách bài"; headers = ["STT", "Tiêu đề", "Mô tả", "Hashtag", "Đường dẫn video"]; widths = [6, 42, 60, 40, 70]; head_fill = PatternFill("solid", fgColor="4F46E5"); head_font = Font(bold=True, color="FFFFFF", size=11); thin = Side(style="thin", color="D9DCE6"); border = Border(left=thin, right=thin, top=thin, bottom=thin); wrap = Alignment(vertical="top", wrap_text=True)
    
    ws.append([f"Series: {_xlsx_text(proj.get("title", ""))}"]); ws["A1"].font = Font(bold=True, size=13, color="1E2B4A"); hr = 2
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.border = border
    
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + c)].width = w
    r = hr + 1
    for i, l in enumerate(lessons, 1):
        full = project_store.get_lesson(project_id, l["id"]) or {}
        sc = full.get("script", {}) or {}
        title = sc.get("title") or l.get("title", "")
        desc = sc.get("description", "") or ""
        subject = sc.get("subject", "") or ""
        tags = make_hashtags(title, subject, proj.get("title", ""))
        path = full.get("rendered_video_path") or full.get("rendered_video_path_9_16") or ""
        vals = [i, _xlsx_text(title), _xlsx_text(desc), _xlsx_text(tags), _xlsx_text(path or "(chưa render)")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = wrap
            cell.border = border
        ws.cell(row=r, column=1).alignment = Alignment(vertical="top", horizontal="center")
        r += 1
    ws.freeze_panes = "A3"
    
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True); wb.save(out_path)
    return out_path
