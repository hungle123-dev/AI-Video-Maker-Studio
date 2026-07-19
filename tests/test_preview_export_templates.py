from __future__ import annotations

import json
import subprocess
import urllib.request
from pathlib import Path

import pytest
from openpyxl import load_workbook
from PIL import Image

from core import backgrounds, preview, preview_server, templates
from core import preview_demo
from core.project_store import ProjectStore
from core.schema import validate_script


def test_preview_server_serves_registered_payload(sample_script):
    token = "pytest-preview"
    timing = preview.build_timing(sample_script)
    preview_server.register(token, sample_script, timing=timing)
    port = preview_server.ensure_server()
    with urllib.request.urlopen(f"http://127.0.0.1:{port}/script/{token}", timeout=5) as response:
        payload = json.loads(response.read().decode("utf-8"))
    assert payload["script"]["title"] == sample_script["title"]
    assert payload["timing"]["total_duration"] > 0

    with urllib.request.urlopen(f"http://127.0.0.1:{port}/preview.html", timeout=5) as response:
        assert b"TubeCraft" in response.read()


def test_actual_static_preview_renderer(tmp_path, sample_script):
    # Registered local scenes remain renderable through the final renderer.
    sample_script["steps"][0]["elements"] = [{
        "type": "custom_js",
        "template": "big_word",
        "params": {"word": "ĐÃ KIỂM TRA", "sub": "Cảnh local"},
    }]
    output = tmp_path / "preview.png"
    project = {
        "theme": "dark",
        "aspect_ratio": "1:1",
        "art_style": "default",
        "subtitle_enabled": False,
    }
    result = preview.render_step_preview(project, sample_script, 0, str(output))
    assert result["ok"], result
    assert output.stat().st_size > 1000


def test_preview_rejects_raw_js_before_node(tmp_path, monkeypatch):
    malicious = {
        "steps": [{"id": 1, "elements": [{
            "type": "custom_js",
            "code": "process.getBuiltinModule('fs').writeFileSync('owned.txt','x');",
        }]}],
    }

    def must_not_spawn(*_args, **_kwargs):
        raise AssertionError("untrusted custom_js reached Node")

    monkeypatch.setattr(preview.subprocess, "run", must_not_spawn)
    result = preview._render(
        "node",
        {},
        {"theme": "dark", "aspect_ratio": "9:16"},
        malicious,
        {"steps": [], "total_duration": 0},
        0.0,
        str(tmp_path / "unsafe.png"),
    )
    assert not result["ok"]
    assert "không an toàn" in result["error"]


def test_preview_rejects_external_image_before_node(tmp_path, monkeypatch):
    malicious = {
        "steps": [{"id": 1, "elements": [{
            "type": "image", "src": "http://127.0.0.1:9/probe.png",
        }]}],
    }

    monkeypatch.setattr(
        preview.subprocess,
        "run",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("external image reached Node")),
    )
    result = preview._render(
        "node", {}, {"theme": "dark", "aspect_ratio": "9:16"}, malicious,
        {"steps": [], "total_duration": 0}, 0.0, str(tmp_path / "unsafe.png"),
    )
    assert not result["ok"]
    assert "không an toàn" in result["error"]


def test_preview_cancellation_terminates_the_node_process(tmp_path, sample_script, monkeypatch):
    class RunningProcess:
        returncode = None

        def __init__(self):
            self.terminated = False

        def poll(self):
            return self.returncode

        def communicate(self, timeout=None):
            raise subprocess.TimeoutExpired("node", timeout)

        def terminate(self):
            self.terminated = True
            self.returncode = -15

        def wait(self, timeout=None):
            return self.returncode

    process = RunningProcess()
    monkeypatch.setattr(preview.subprocess, "Popen", lambda *_args, **_kwargs: process)
    checks = iter((False, True))
    result = preview._render(
        "node", {}, {"theme": "dark", "aspect_ratio": "9:16"}, sample_script,
        preview.build_timing(sample_script), 1.0, str(tmp_path / "cancelled.png"),
        cancel_check=lambda: next(checks),
    )
    assert not result["ok"] and result.get("cancelled")
    assert process.terminated


def test_final_preview_loads_only_a_real_local_gallery_asset(tmp_path, monkeypatch):
    import config

    gallery = tmp_path / "gallery"
    gallery.mkdir()
    asset = gallery / "logo.png"
    Image.new("RGB", (32, 20), color=(30, 160, 220)).save(asset)
    monkeypatch.setattr(config, "GALLERY_DIR", gallery)
    script = {"steps": [{"id": 1, "voice_text": "Ảnh local", "elements": [{
        "type": "image", "src": "gallery:logo.png", "width": 160, "height": 100,
    }]}]}
    output = tmp_path / "gallery-preview.png"
    result = preview.render_step_preview(
        {"theme": "dark", "aspect_ratio": "1:1", "subtitle_enabled": False},
        script,
        0,
        str(output),
    )
    assert result["ok"], result
    assert output.stat().st_size > 1_000


def test_final_preview_fails_when_a_declared_gallery_asset_is_missing(tmp_path, monkeypatch):
    import config

    gallery = tmp_path / "empty-gallery"
    gallery.mkdir()
    monkeypatch.setattr(config, "GALLERY_DIR", gallery)
    script = {"steps": [{"id": 1, "voice_text": "Ảnh local", "elements": [{
        "type": "image", "src": "gallery:missing.png", "width": 160, "height": 100,
    }]}]}
    result = preview.render_step_preview(
        {"theme": "dark", "aspect_ratio": "1:1", "subtitle_enabled": False},
        script,
        0,
        str(tmp_path / "missing.png"),
    )
    assert not result["ok"]
    assert "image_asset_error" in result["error"]


def test_template_customization_and_backgrounds(tmp_path, monkeypatch):
    store_path = tmp_path / "templates.json"
    monkeypatch.setattr(templates, "_store_path", lambda: store_path)
    monkeypatch.setattr(templates, "_thumb_cache_dir", lambda: tmp_path / "thumb-cache")
    store_path.write_text("[]", encoding="utf-8")
    assert templates.get_template("lux_finance")["id"] == "lux_finance"
    changed = templates.update_template("lux_finance", title_color="#123456")
    assert changed["title_color"] == "#123456"
    assert templates.is_customized("lux_finance")
    assert templates.thumb_path("lux_finance").parent == tmp_path / "thumb-cache"
    reset = templates.reset_template("lux_finance")
    assert reset["title_color"] == "#FFD700"
    assert len(templates.list_templates()) >= 21
    assert backgrounds.get("midnight")["grad"] == ["#0a0a1a", "#1a1030"]
    assert backgrounds.contrast_warning("white_flat", "default")


def test_template_exemplar_preview_is_declarative_local_scene():
    script, errors = validate_script(preview_demo.demo_script("light_news"))
    assert not errors
    visual = next(
        item for item in script["steps"][0]["elements"]
        if item.get("type") == "custom_js"
    )
    assert visual.get("template")
    assert visual["trusted_template"] == visual["template"]


@pytest.mark.parametrize("template_id", ["lux_finance", "light_news", "math_noir"])
def test_template_preview_smoke_uses_final_renderer(tmp_path, template_id):
    output = tmp_path / f"{template_id}.png"
    result = preview_demo.render_preview_png(template_id, str(output), seconds=0.4)
    assert result == str(output)
    assert output.stat().st_size > 1_000


def test_excel_export_workflow(tmp_path, monkeypatch, script_copy):
    store = ProjectStore(tmp_path / "projects")
    project = store.create_project("Series Excel")
    lesson = store.create_lesson(project["id"], "Bài Excel")
    store.save_script(project["id"], lesson["id"], script_copy)
    import core.project_store as project_store_module

    monkeypatch.setattr(project_store_module, "project_store", store)
    from core.export_excel import export_project_xlsx

    output = Path(export_project_xlsx(project["id"], str(tmp_path / "export.xlsx")))
    workbook = load_workbook(output)
    sheet = workbook["Danh sách bài"]
    assert sheet["A1"].value == "Series: Series Excel"
    assert sheet["B3"].value == script_copy["title"]
    assert "#" in sheet["D3"].value


def test_excel_export_escapes_formula_like_user_content(tmp_path, monkeypatch, script_copy):
    store = ProjectStore(tmp_path / "projects")
    project = store.create_project("Formula safety")
    lesson = store.create_lesson(project["id"], "Bài Excel")
    unsafe = dict(script_copy, title="=1+1", description="@HYPERLINK(\"https://example.invalid\")")
    assert store.save_script(project["id"], lesson["id"], unsafe)[0]
    import core.project_store as project_store_module

    monkeypatch.setattr(project_store_module, "project_store", store)
    from core.export_excel import export_project_xlsx

    workbook = load_workbook(export_project_xlsx(project["id"], str(tmp_path / "formula.xlsx")))
    sheet = workbook["Danh sách bài"]
    assert sheet["B3"].data_type == "s"
    assert sheet["C3"].data_type == "s"
    assert sheet["B3"].value == "'=1+1"
    assert sheet["C3"].value.startswith("'@HYPERLINK")


def test_excel_export_never_overwrites_an_existing_path(tmp_path, monkeypatch, script_copy):
    store = ProjectStore(tmp_path / "projects")
    project = store.create_project("No overwrite")
    lesson = store.create_lesson(project["id"], "Bài Excel")
    assert store.save_script(project["id"], lesson["id"], script_copy)[0]
    import core.project_store as project_store_module

    monkeypatch.setattr(project_store_module, "project_store", store)
    from core.export_excel import export_project_xlsx

    requested = tmp_path / "export.xlsx"
    requested.write_bytes(b"keep-this-export")
    created = Path(export_project_xlsx(project["id"], str(requested)))
    assert requested.read_bytes() == b"keep-this-export"
    assert created != requested and created.is_file()
